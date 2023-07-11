from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import time
import openpyxl

res = []
path = "C:\\Users\\nballa\\Desktop\\python\\chrome1.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
for i in range(1, m_row + 1):
	k = sheet_obj.cell(row = i, column = 1)
	#m = "('malware' OR 'virus' OR 'adware' OR 'exploit' OR 'vulnerabilities')"
	z = k.value

	driver = webdriver.Chrome('C:\\Users\\chromedriver.exe')

	driver.get('https://www.google.com')

	search = driver.find_element("name",'q')
	search.send_keys(z)
	time.sleep(5)
	search.send_keys(Keys.RETURN)
	time.sleep(2)
	#search_btn = driver.find_element_by_css_selector('input[type="submit"]')
	#search_btn.click()
	try:
		main = WebDriverWait(driver, 10).until(
			EC.presence_of_element_located((By.ID, "search"))
		)
		link = main.find_element(By.TAG_NAME, 'a')
		print(link.get_attribute("href"))
		res.append([z, link.get_attribute("href")])
	
	except NoSuchElementException:
		print("No link found for {}; outputting 'null'".format(z))
		res.append([z, "null"])

	finally:
	    driver.quit()
wb = openpyxl.Workbook()
sheet = wb.active
sheet_title = "chromeextensionvalidation"
for i in res:
	sheet.append(i)
wb.save("C:\\Users\\nballa\\Desktop\\python\\chromeextensionvalidation2.xlsx")
